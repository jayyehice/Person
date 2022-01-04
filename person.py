# -*- coding: utf-8 -*-
"""
Created on Thu Mar  4 21:20:40 2021

@author: 崇杰
"""

import pypyodbc
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border,Side

def Draw_Frame(ws, LT, RB):#LT=左上格編號,RB=右下格編號
    column1=[ord(LT[0]), int(LT[1:])]
    column2=[ord(RB[0]), int(RB[1:])]
    side=Side(border_style='thin',color='000000')
    
    if LT[0] != RB[0] and LT[1:] != RB[1:]:
        ws[LT].border=Border(top=side, left=side)#左上
        ws[LT[0]+RB[1:]].border=Border(bottom=side, left=side)#左下
        ws[RB[0]+LT[1:]].border=Border(top=side, right=side)#右上
        ws[RB].border=Border(bottom=side, right=side)#右下
        #畫上方框線
        for i in range(column1[0]+1,column2[0]):
            alphabet = chr(i)
            ws[alphabet+LT[1:]].border=Border(top=side)
            
        
        #畫下方框線
        for j in range(column1[0]+1,column2[0]):
            alphabet = chr(j)
            ws[alphabet+RB[1:]].border=Border(bottom=side)
            
        
        #畫左方框線
        for k in range(column1[1]+1,column2[1]):
            number = str(k)
            ws[LT[0]+number].border=Border(left=side)    
            
        #畫右方框線
        for l in range(column1[1]+1,column2[1]):
            number = str(l)
            ws[RB[0]+number].border=Border(right=side)
    
    elif LT[0] != RB[0] and LT[1:] == RB[1:]:
        ws[LT].border=Border(top=side, bottom=side, left=side)#左
        ws[RB].border=Border(top=side, bottom=side, right=side)#右
        #畫上下框線
        for i in range(column1[0]+1,column2[0]):
            alphabet = chr(i)
            ws[alphabet+LT[1:]].border=Border(top=side, bottom=side)

    elif LT[0] == RB[0] and LT[1:] != RB[1:]:
        ws[LT].border=Border(top=side, right=side, left=side)#上
        ws[RB].border=Border(bottom=side, right=side, left=side)#下
        #畫左右框線
        for k in range(column1[1]+1,column2[1]):
            number = str(k)
            ws[LT[0]+number].border=Border(right=side, left=side)    
    
    else:#單一格
        ws[LT].border=Border(top=side, bottom=side, right=side, left=side)
        
#------------------------------------------------------------------------------ 

db_file = './2021db.mdb' ## Microsoft Access 檔案名稱
user = ''
password = ''
connection_string = 'DRIVER={Microsoft Access Driver (*.mdb)};DBQ=%s;UID=%s;PWD=%s' % (db_file, user, password)
conn = pypyodbc.win_connect_mdb(connection_string)
cur = conn.cursor()

data = []#用來儲存資料

subject = [['月定獻金', '4101%'], ['聖餐獻金', '4103%'], ['節期獻金', '4104%'],
        ['感恩獻金', '4105%'], ['特別獻金', '4106%'], ['建築及專案獻金', '4211.01'],
        ['對外獻金-本宗', '4214%'], ['對內獻金', '4215%']]





Fellowship = ['123']



SQL = "SELECT id, name FROM [member] ORDER BY id;"
cur.execute(SQL)
member_list = cur.fetchall()


for name in member_list:
    

    print(name[0], name[1]) 
   
    #輸入SQL語法
    #SQL = "SELECT SUM(amt) FROM [per] WHERE name = '"+ name +"' and code_des = '月定獻金';" ## invo為資料表名稱
    for subj in subject:
        
        SQL = "SELECT SUM(amt) FROM [per] WHERE name = '"+ name[1] +"' and code LIKE '"+subj[1]+"';"
        cur.execute(SQL)
        list_ = cur.fetchall()
        print(subj[0], list_[0][0])
    #print(list_)


'''    
    #若沒資料就跳過不顯示
    if list_ == []:
        continue
        
    for i in range(len(list_)):
        temp = []
        time = str(list_[i][1])
        money = int(list_[i][3])
        if i == 0:
            temp=[name[0]]
            data.append(temp)
            
        temp = ['', time[:10], list_[i][2], list_[i][0], money]
        data.append(temp)
    
    data.append('sum_mark')#事工單位結束標誌


'''
cur.close()
conn.close()

#------------------------------------------------------------------------------ 
       
'''        
#抓今天列印日期
print_day = str(datetime.date.today())
p_day = print_day.replace('-', '')


# 建立新工作簿
wb = Workbook()

# 建立新工作表 data
ws1 = wb.create_sheet("各事工單位", 0)



#收入頁首
ws1.merge_cells('A2:J2')
ws1.merge_cells('A3:J3')
ws1['A2']='台灣基督長老教會七星中會吳興教會'
ws1['A3']= '各事工單位收支表'
ws1['A2'].alignment = Alignment(horizontal="center", vertical="center")
ws1['A3'].alignment = Alignment(horizontal="center", vertical="center")
ws1['A2'].font = Font( size=16, bold=True )
ws1['A3'].font = Font( bold=True )


#英文欄位列表
row_list=['B', 'D', 'E']

row = 4

income = 0#收入小計
cost = 0#支出小計


for i in data:
    row += 1
    k = 0 #計數用
    if i != 'sum_mark':
        if i[0] != '':
            f = str(row+1)#畫外框第一格紀錄
            #表頭
            ws1['B'+str(row)]=i[0]
            ws1['D'+str(row)]='日期'
            ws1['E'+str(row)]='摘要'
            ws1['H'+str(row)]='收入'
            ws1['I'+str(row)]='支出'
            Draw_Frame(ws1, 'B'+str(row), 'I'+str(row))
        else:
            for j in row_list:
                if i[k] == None:
                    i[k] = ''
                if k == 1:
                    i[k] = i[k].replace('2021-','')
                    i[k] = i[k].replace('-','/')
                ws1[j+str(row)]=i[k]
                k += 1
            
            if i[k] == '收入':
                ws1['H'+str(row)]=i[k+1]
                ws1['H'+str(row)].number_format = '$#,##'
                income += i[k+1]
            else:
                ws1['I'+str(row)]=i[k+1]
                ws1['I'+str(row)].number_format = '$#,##'
                cost += i[k+1]
    else:
        row += 1
        ws1['G'+str(row)]='小計'
        ws1['G'+str(row)].font = Font( bold=True )
        if income != 0:
            ws1['H'+str(row)]=income
            ws1['H'+str(row)].number_format = '$#,##'
            ws1['H'+str(row)].font = Font( bold=True )
        if cost != 0:
            ws1['I'+str(row)]=cost
            ws1['I'+str(row)].number_format = '$#,##'
            ws1['I'+str(row)].font = Font( bold=True )
        Draw_Frame(ws1, 'B'+f, 'I'+str(row))
        row += 2
        income = 0
        cost = 0

# 保存文件

file_name = './吳興教會收支表-' + p_day + '.xlsx'
wb.save(r''+file_name)

'''